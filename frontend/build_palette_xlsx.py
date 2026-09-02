#!/usr/bin/env python3
"""Build a readable Excel palette sheet for the LibExpress frontend concept."""

import shutil
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter

OUT = "/Users/matthewestilo/Documents/GitHub/LibExpress/frontend/LibExpress-Palette.xlsx"

# ---------------------------------------------------------------------------
# palette data: (group, token, hex_or_rgba, rgb_tuple_or_None, usage)
# rgb_tuple is (r,g,b) 0-255 when we have a solid hex; None for translucent
# ---------------------------------------------------------------------------
GROUPS = [
    ("BACKGROUND & SURFACES", [
        ("--bg",         "#0a0a0f", (10, 10, 15),   "Root / page background (near-black)"),
        ("--panel",      "#111118", (17, 17, 24),   "Card / panel surface"),
        ("--panel-2",    "#16161f", (22, 22, 31),   "Slightly lighter nested panel / hover state"),
        ("--border",     "#1e1e2a", (30, 30, 42),   "Subtle borders (tables, dividers)"),
        ("--border-2",   "#2a2a3a", (42, 42, 58),   "Secondary / hover borders"),
    ]),
    ("TEXT", [
        ("--text",       "#e8e8f0", (232, 232, 240), "Primary text (off-white on dark)"),
        ("--muted",      "#6b6b80", (107, 107, 128), "Secondary / muted text"),
    ]),
    ("NEON ACCENTS", [
        ("--neon-blue",  "#00d4ff", (0, 212, 255),   "Neon blue — ACTIVE status, low-stock, primary accent"),
        ("--neon-green", "#39ff14", (57, 255, 20),   "Neon green — RETURNED, available, PAID, ACTIVE member"),
    ]),
    ("ALERT / STATUS", [
        ("--alert",      "#ff5e5e", (255, 94, 94),  "Muted red — OVERDUE, unpaid, SUSPENDED (used sparingly)"),
    ]),
    ("TRANSLUCENT / OVERLAY VARIANTS", [
        ("--neon-blue-dim",  "rgba(0,212,255,0.12)", None, "Subtle blue backdrop for active tabs / pills"),
        ("--neon-green-dim", "rgba(57,255,20,0.10)", None, "Subtle green backdrop for success pills"),
        ("--alert-dim",      "rgba(255,94,94,0.10)", None, "Subtle red backdrop for overdue / unpaid rows"),
        ("--shadow",         "rgba(0,0,0,0.5)",      None, "Drop shadow for panels"),
    ]),
]

# ---------------------------------------------------------------------------
# styling helpers
# ---------------------------------------------------------------------------
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
WRAP   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

WHITE_BOLD = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
DARK_BOLD  = Font(name="Calibri", bold=True, color="0A0A0F", size=11)
DARK_REG   = Font(name="Calibri", color="0A0A0F", size=11)
LIGHT_REG  = Font(name="Calibri", color="6B6B80", size=10)

def hex_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color.replace("#", ""), end_color=hex_color.replace("#", ""), fill_type="solid")

wb = Workbook()
ws = wb.active
ws.title = "LibExpress Palette"

# column widths
widths = {"A": 26, "B": 14, "C": 14, "D": 56, "E": 16}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

row = 1

# ---- Title block ----
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
c = ws.cell(row=row, column=1, value="LibExpress — Frontend Concept Color Palette")
c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
c.fill = PatternFill(start_color="0A0A0F", end_color="0A0A0F", fill_type="solid")
c.alignment = CENTER
ws.row_dimensions[row].height = 30
row += 1

ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
c = ws.cell(row=row, column=1, value="Dark chrome UI · neon blue #00d4ff · neon green #39ff14 · sparing alert #ff5e5e")
c.font = Font(name="Calibri", italic=True, color="6B6B80", size=10)
c.alignment = CENTER
ws.row_dimensions[row].height = 18
row += 2

# ---- header row ----
headers = ["Token", "Hex", "RGB", "Usage / Role", "Swatch"]
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=row, column=i, value=h)
    c.font = WHITE_BOLD
    c.fill = PatternFill(start_color="2A2A3A", end_color="2A2A3A", fill_type="solid")
    c.alignment = CENTER
    c.border = BORDER
ws.row_dimensions[row].height = 20
row += 1

# ---- data rows grouped by section ----
for group_name, entries in GROUPS:
    # group divider row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws.cell(row=row, column=1, value=group_name)
    c.font = DARK_BOLD
    c.fill = PatternFill(start_color="16161F", end_color="16161F", fill_type="solid")
    c.alignment = LEFT
    c.border = BORDER
    ws.row_dimensions[row].height = 20
    row += 1

    for token, hexval, rgb, usage in entries:
        ws.cell(row=row, column=1, value=token).font = DARK_BOLD
        ws.cell(row=row, column=1).alignment = LEFT
        ws.cell(row=row, column=1).border = BORDER

        c_hex = ws.cell(row=row, column=2, value=hexval)
        c_hex.font = DARK_REG
        c_hex.alignment = CENTER
        c_hex.border = BORDER
        c_hex.number_format = "@"  # text, keep leading zeros friendly (no # in these though)

        if rgb is not None:
            r, g, b = rgb
            c_rgb = ws.cell(row=row, column=3, value=f"{r}, {g}, {b}")
        else:
            c_rgb = ws.cell(row=row, column=3, value="—")
        c_rgb.font = LIGHT_REG
        c_rgb.alignment = CENTER
        c_rgb.border = BORDER

        c_use = ws.cell(row=row, column=4, value=usage)
        c_use.font = DARK_REG
        c_use.alignment = WRAP
        c_use.border = BORDER

        # swatch: fill with the actual color; for rgba, fill with a light tint approximation
        if rgb is not None:
            swatch = hex_fill(hexval)
        else:
            # translucent variants: put a light neutral tint with the hex string as value so it's still readable
            swatch = PatternFill(start_color="E8E8F0", end_color="E8E8F0", fill_type="solid")
        c_sw = ws.cell(row=row, column=5)
        c_sw.fill = swatch
        c_sw.border = BORDER
        # put the hex in the swatch cell as tiny text so inspectors can copy it
        c_sw.value = hexval if rgb is None else ""
        if rgb is None:
            c_sw.font = Font(name="Calibri", color="6B6B80", size=9)
            c_sw.alignment = CENTER

        ws.row_dimensions[row].height = 18
        row += 1

    row += 1  # space between groups

# ---- footer note ----
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
c = ws.cell(row=row, column=1, value="Source: frontend/src/index.css (vars defined in :root)")
c.font = Font(name="Calibri", italic=True, color="6B6B80", size=9)
c.alignment = Alignment(horizontal="right", vertical="center")

# ---------------------------------------------------------------------------
# print setup: fit to width, light gridlines off
# ---------------------------------------------------------------------------
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.sheet_view.showGridLines = False

wb.save(OUT)
print(f"Wrote {OUT}")
