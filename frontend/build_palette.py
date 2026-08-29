import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

OUT = "/Users/matthewestilo/Documents/GitHub/LibExpress/frontend/LibExpress-Palette.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "LibExpress Palette"

header_fill = PatternFill("solid", fgColor="111118")
header_font = Font(color="E8E8F0", bold=True, size=11)
group_fill  = PatternFill("solid", fgColor="1E1E2A")
group_font  = Font(color="00D4FF", bold=True, size=11)
data_font   = Font(color="E8E8F0", size=10)
rgba_font   = Font(color="6B6B80", italic=True, size=10)
thin = Side(style="thin", color="3A3A4A")
cell_border = Border(bottom=thin)

widths = {"A": 22, "B": 12, "C": 14, "D": 54, "E": 14}
for col_name, w in widths.items():
    ws.column_dimensions[col_name].width = w

row = 1
ws.merge_cells("A1:E1")
ws["A1"] = "LibExpress — Frontend Concept Color Palette"
ws["A1"].font = Font(color="E8E8F0", bold=True, size=14)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 26

row = 2
ws.merge_cells("A2:E2")
ws["A2"] = "Dark chrome UI · Neon blue #00D4FF · Neon green #39FF14 · Sparing alert #FF5E5E"
ws["A2"].font = Font(color="6B6B80", italic=True, size=10)
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 16

row = 4
for col_index, header in enumerate(["Token", "Hex", "RGB", "Usage / Role", "Swatch"], start=1):
    c = ws.cell(row=row, column=col_index, value=header)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(bottom=Side(style="medium", color="2A2A3A"))
ws.row_dimensions[row].height = 20
row += 1

groups = [
    ("BACKGROUND & SURFACES", [
        ("--bg",         "#0A0A0F", "10, 10, 15",   "Root / page background (near-black)"),
        ("--panel",      "#111118", "17, 17, 24",   "Card / panel surface"),
        ("--panel-2",    "#16161F", "22, 22, 31",   "Slightly lighter nested panel / hover state"),
        ("--border",     "#1E1E2A", "30, 30, 42",   "Subtle borders (tables, dividers)"),
        ("--border-2",   "#2A2A3A", "42, 42, 58",   "Secondary / hover borders"),
    ]),
    ("TEXT", [
        ("--text",       "#E8E8F0", "232, 232, 240", "Primary text (off-white on dark)"),
        ("--muted",      "#6B6B80", "107, 107, 128", "Secondary / muted text"),
    ]),
    ("NEON ACCENTS", [
        ("--neon-blue",  "#00D4FF", "0, 212, 255",   "Neon blue — ACTIVE status, low-stock, primary accent"),
        ("--neon-green", "#39FF14", "57, 255, 20",    "Neon green — RETURNED, available, PAID, ACTIVE member"),
    ]),
    ("ALERT / STATUS", [
        ("--alert",      "#FF5E5E", "255, 94, 94",   "Muted red — OVERDUE, unpaid, SUSPENDED (used sparingly)"),
    ]),
    ("TRANSLUCENT / OVERLAY VARIANTS", [
        ("--neon-blue-dim",  "rgba(0,212,255,0.12)", "—", "Subtle blue backdrop for active tabs / pills"),
        ("--neon-green-dim", "rgba(57,255,20,0.10)", "—", "Subtle green backdrop for success pills"),
        ("--alert-dim",      "rgba(255,94,94,0.10)", "—", "Subtle red backdrop for overdue / unpaid rows"),
        ("--shadow",         "rgba(0,0,0,0.5)",      "—", "Drop shadow for panels"),
    ]),
]

for gname, entries in groups:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws.cell(row=row, column=1, value=gname)
    c.fill = group_fill
    c.font = group_font
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18
    row += 1
    for token, hexval, rgb, usage in entries:
        ws.cell(row=row, column=1, value=token).font = Font(color="E8E8F0", bold=True, size=10)
        ws.cell(row=row, column=2, value=hexval).font = Font(color="E8E8F0", size=10)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=rgb).font = rgba_font if rgb == "—" else data_font
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=usage).font = data_font
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        swatch = ws.cell(row=row, column=5)
        if hexval.startswith("#"):
            swatch.fill = PatternFill("solid", fgColor=hexval.lstrip("#"))
        else:
            swatch.fill = PatternFill("solid", fgColor="E8E8F0")
            swatch.value = hexval
            swatch.font = rgba_font
            swatch.alignment = Alignment(horizontal="center")
        swatch.border = cell_border
        for col_idx in range(1, 5):
            ws.cell(row=row, column=col_idx).border = cell_border
        ws.row_dimensions[row].height = 16
        row += 1
    row += 1

row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
note = ws.cell(row=row, column=1, value="Source: frontend/src/index.css (:root CSS variables)")
note.font = Font(color="6B6B80", italic=True, size=9)
note.alignment = Alignment(horizontal="right")

ws.sheet_view.showGridLines = False
ws.freeze_panes = "A5"

wb.save(OUT)
print(f"Wrote {OUT}")
